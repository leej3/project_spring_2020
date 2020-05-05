import pytest
from eeg_gsr_wellness_tracker import append_table
#from eeg_gsr_wellness_tracker import Mood_Focus_Table
import pandas
import mne
import numpy
from pathlib import Path
from openpyxl import load_workbook
workbook = load_workbook(filename='Mood_Focus_Table.xlsx')
workbook.sheetnames


'''
def extract_mean_amplitude_to_data():
   # assert append_table.extract_mean_amplitude_to_data(Picks = 'EEG Fp2-LE') type(sumdf) is pandas.core.frame.DataFrame
    assert append_table.extract_mean_amplitude_to_data(Picks = 'EEG Fp2-LE') type(raw) is mne.io.edf.edf.RawEDF
    assert append_table.extract_mean_amplitude_to_data(Picks = 'EEG Fp2-LE') type("time") is numpy.int64
'''



#def read_raw_eeg_file(eeg_file_path):
#    assert isinstance(eeg_file_path,  str)
#    assert isinstance(raw,  RawEDF)
    # you should add some small data in the tests directory for this.
    #filename = ‘some_data.edf’
    
    
def test_read_raw_eeg_file():
    file = 'AusEC.edf'
    filename = 'AusEC.edf'
    first_3_values = [-1.21509270e-05, -1.13508659e-05, -1.09508354e-05]
    #first_value = ['4.62613']
    data = mne.io.read_raw_edf(file)
    raw_data = data.get_data()
    assert first_3_values == raw_data
     # you should add some small data in the tests directory for this.
   # filename = ‘some_data.edf’
   # first_ten_values = [1,2,3,2,2,0,0,0,-10,20]
   # result = append_table.read_raw_eeg_file(filename)
   # assert first_ten_values == result[:10]
    
    
    
    
    
    
    
    
@pytest.mark.parametrize("Picks", ["EEG F3-LE"])
@pytest.mark.parametrize("excel_file", ["Mood_Focus_Table.xlsx"])
def test_extract_mean_amplitude_to_data(Picks, excel_file):
    #file_dir = excel_file / 'file'
    #extract_mean_amplitude_to_data(Picks = 'EEG Fp2-LE')
    assert isinstance(Picks,  str)
    #assert isinstance(sumdf, )
    assert type("time") is str
    # assert MeanPzdf != 0
    # assert MeanPzdf == Pzdf.mean()
    #assert picks==Picks
    #assert Path(excel_file).is_file()
    #assert isin(df.columns)
    #assert isinstance (Picks, channel) Pick.type() == "channel
   
    #assert excel_file.is_file()
    #filename = 'Mood_Focus_Table.xlsx'
    #first_ten_values = [1,2,3,2,2,0,0,0,-10,20]   
    #filename = 'Mood_Focus_Table.xlsx'
    #first_value = ['4.62613']
    #sheet_value = sheet.cell(row=27, column=5).value
    #assert first_value == sheet_value